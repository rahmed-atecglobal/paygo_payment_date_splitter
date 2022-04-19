import numpy as np
import os
import openpyxl
from openpyxl.styles import Font
import pandas as pd

'''
in_wb = openpyxl.load_workbook('input/angaza_payment.xlsx')
in_sheet = in_wb['ap']

max_row = in_sheet.max_row
max_col = in_sheet.max_column

row_no = 2

print(in_sheet)
'''

#wb = openpyxl.load_workbook('output/payment_info.xlsx')
wb = openpyxl.Workbook()
ws = wb.active

df = pd.read_excel('input/angaza_payment.xlsx')
#print(df['E'].unique())
#print(df)
dlist1 = df['Account Number'].unique()
nan_list = np.isnan(dlist1)
not_nan_list = ~nan_list
dlist = dlist1[not_nan_list]
dlist = np.sort(dlist)
counter = 1

row_no = 2

max_no_date = 1

for payid in dlist:
    #print(i)
    row_list = df.index[df['Account Number']==payid].tolist()
    product_name = df['Group Name'][row_list[0]]

    ws['A' + str(row_no)] = str(payid)[:-2]
    ws['B' + str(row_no)] = product_name

    date_list = []
    for rows in row_list:
        date = df['Recorded (UTC)'][rows]
        date_list.append(date)
    #print(f"{counter}. {payid} :: {product_name}  :: {row_list} : {date_list}")

    if len(date_list) > max_no_date:
        max_no_date = len(date_list)

    ws['C' + str(row_no)] = len(date_list)
    col_counter = 4
    date_list.reverse()
    for date in date_list:
        cell_data  = ws.cell(row=row_no, column=col_counter)
        col_counter += 1
        dst = str(date).split()
        cell_data.value = dst[0]
    counter += 1
    row_no += 1

ws['A1'] = "PayGo ID"
ws['B1'] = "Payment Group"
ws['C1'] = "Total times Paid"

for i in range(1,max_no_date+1):
    t_cell = ws.cell(row=1, column=3+i)
    t_cell.value = "Date " + str(i)

wb.save('output/payment_info.xlsx')


